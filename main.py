import sys
import os
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QFileDialog
from PyQt5 import uic

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    class Workbook: pass
    class PatternFill:
        def __init__(self, *args, **kwargs): pass
    class Font:
        def __init__(self, *args, **kwargs): pass
    class Alignment:
        def __init__(self, *args, **kwargs): pass
    print("ADVERTENCIA: La librería 'openpyxl' no está instalada. El exportador XLSX no funcionará.")

# Importaciones del proyecto
from database.db_conexion import DatabaseManager
from controllers.gestion_datos import GestionDatosController
from controllers.vista_horario import VistaHorarioController

try:
    from controllers.generador import GeneradorController
except ImportError:
    class GeneradorController:
        def __init__(self, ui, db, main_window=None): pass


class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # Carga la parte de la izquierda que siempre se mantiene
        ui_path = os.path.join(os.path.dirname(__file__), 'ui', 'main_shell.ui')
        uic.loadUi(ui_path, self)
        
        # Inicia la conexion con la BBDD
        self.db = DatabaseManager()
        
        # Conectamos botones de la parte de la izquierda
        self.btnVistaHorario.clicked.connect(self.cargar_vista_horario)
        self.btnGestionDatos.clicked.connect(self.cargar_gestion_datos)
        self.btnGenerarHorario.clicked.connect(self.cargar_generador)
        self.btnExportar.clicked.connect(self.exportar_xlsx)
        
        self.cargar_vista_horario()

        # Cargar primera página
        self.btnVistaHorario.click()

    #Gestion de paginas UI

    #Metodo para limpiar el contenedor donde se cargan las paginas
    def limpiar_contenedor(self):
        """Elimina el widget actual del layout principal"""
        layout = self.layout_contenedor
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    #Metodo para cargar una pagina UI en el contenedor principal
    def cargar_ui_pagina(self, nombre_archivo):
        
        self.limpiar_contenedor()
        
        path = os.path.join(os.path.dirname(__file__), 'ui', 'pages', nombre_archivo)
        try:
            nuevo_widget = uic.loadUi(path)
            self.layout_contenedor.addWidget(nuevo_widget)
            return nuevo_widget
        except FileNotFoundError:
            print(f"No se pudo encontrar el archivo: {nombre_archivo}")
            return None

    #Funciones para cargar cada pagina UI
    def cargar_vista_horario(self):
        widget = self.cargar_ui_pagina('pagina_horario.ui')
        if widget:
            self.ctrl_vista = VistaHorarioController(widget, self.db, self)
            self.resaltar_boton(self.btnVistaHorario)

    def cargar_gestion_datos(self):
        widget = self.cargar_ui_pagina('pagina_gestion.ui')
        if widget:
            self.ctrl_gestion = GestionDatosController(widget, self.db)
            self.ctrl_gestion.cargar_datos_iniciales()
            self.resaltar_boton(self.btnGestionDatos)

    def cargar_generador(self):
        widget = self.cargar_ui_pagina('pagina_generador.ui')
        if widget:
            self.ctrl_gen = GeneradorController(widget, self.db, self)
            self.resaltar_boton(self.btnGenerarHorario)

    #Exportar XLSX de datos y horario en hojas separadas y con colores
    def exportar_xlsx(self):

        if not self.db:
            QMessageBox.critical(self, "Error", "La conexión a la base de datos no está disponible.")
            return
            
        if 'Workbook' not in globals() or not isinstance(Workbook, type):
            QMessageBox.critical(self, "Error", "No se encontró la librería 'openpyxl'. Por favor, instálala con: pip install openpyxl")
            return

        # Para pedir ubicación de guardado
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Horario", "datos_horario", "Archivos Excel (*.xlsx)")
        if not path:
            return
        
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'

        try:
            # Crear un nuevo libro de trabajo de OpenPyXL
            wb = Workbook()
            
            # Hoja 1: Datos de gestión
            ws_gestion = wb.active 
            ws_gestion.title = "Datos de Gestion"

            profes = self.db.obtener_profesores()
            modulos = self.db.obtener_modulos()
            prefs = self.db.obtener_todas_preferencias()
            
            # Primer bloque: Profesores y colores
            ws_gestion.append(["--- PROFESORES Y COLORES ---"])
            if profes:
                # Cabeceras
                keys = list(profes[0].keys())
                ws_gestion.append([k.replace('_', ' ').title() for k in keys])
                
                for p in profes:
                    row_data = [p.get(k) for k in keys]
                    ws_gestion.append(row_data)
                    
                    # Colorear la fila
                    color_hex = p.get('color_asignado')
                    if color_hex and len(color_hex) == 7 and color_hex.startswith('#'):
                        try:
                            fill = PatternFill(start_color=color_hex[1:], end_color=color_hex[1:], fill_type="solid")
                            for cell in ws_gestion[ws_gestion.max_row]:
                                cell.fill = fill
                        except Exception as e:
                            pass
                            
            # Segundo bloque: Módulos
            ws_gestion.append([])
            ws_gestion.append(["--- MÓDULOS ---"])
            if modulos:
                keys = list(modulos[0].keys())
                ws_gestion.append([k.replace('_', ' ').title() for k in keys])
                for m in modulos:
                    ws_gestion.append([m.get(k) for k in keys])

            # Tercer bloque: Preferencias
            ws_gestion.append([])
            ws_gestion.append(["--- PREFERENCIAS Y BLOQUEOS ---"])
            if prefs:
                keys = list(prefs[0].keys())
                ws_gestion.append([k.replace('_', ' ').title() for k in keys])
                for pr in prefs:
                    ws_gestion.append([pr.get(k) for k in keys])
                    
            # Hoja 2: Horario generado
            ws_horario = wb.create_sheet(title="Horario Completo")
            horario_datos = self.db.obtener_horario_completo_para_exportar()
            
            if horario_datos:
                # Mapeo de franjas horarias a números
                franja_map = {
                    1: "08:30 - 09:25",
                    2: "09:25 - 10:20",
                    3: "10:20 - 11:15",
                    4: "11:45 - 12:40",
                    5: "12:40 - 13:35",
                    6: "13:35 - 14:30"
                }
                
                # Crear estructura visual: días en columnas, horas en filas
                dias_orden = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]
                horas = list(franja_map.values())
                
                # Encabezados: HORA | LUNES | MARTES | MIÉRCOLES | JUEVES | VIERNES
                ws_horario.append(["HORA"] + dias_orden)
                
                # Establecer estilos para encabezados
                header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
                header_font = Font(bold=True, color="94a3b8")
                for col in range(1, len(dias_orden) + 2):
                    cell = ws_horario.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Agrupar datos por día y franja horaria
                horario_dict = {}
                for h in horario_datos:
                    dia = h.get('Día', '').upper().replace('É', 'E')
                    # Normalizar nombre del día
                    for d in dias_orden:
                        if d.replace('É', 'E') == dia:
                            dia = d
                            break
                    
                    franja = h.get('Franja Horaria', 1)
                    if isinstance(franja, str):
                        try:
                            franja = int(franja)
                        except:
                            franja = 1
                    
                    if dia not in horario_dict:
                        horario_dict[dia] = {}
                    
                    if franja not in horario_dict[dia]:
                        horario_dict[dia][franja] = []
                    
                    horario_dict[dia][franja].append(h)
                
                # Rellenar filas (una fila por hora)
                for franja_num, hora in enumerate(horas, 1):
                    row_data = [hora]
                    row_colors = ["0f172a"]  # Color para columna de horas
                    
                    for dia in dias_orden:
                        if dia in horario_dict and franja_num in horario_dict[dia]:
                            items = horario_dict[dia][franja_num]
                            # Concatenar información de módulos/profesores
                            textos = []
                            for item in items:
                                modulo = item.get('Nombre Módulo', 'N/A')
                                profesor = item.get('Nombre Profesor', '')
                                textos.append(f"{modulo}\n{profesor}")
                            texto = "\n".join(textos)
                            row_data.append(texto)
                            # Usar el color del primer item
                            color = items[0].get('Color Asignado', '#1e293b')
                            row_colors.append(color)
                        else:
                            row_data.append("")
                            row_colors.append("1e293b")
                    
                    ws_horario.append(row_data)
                    
                    # Aplicar colores y estilos a la fila
                    for col, color_hex in enumerate(row_colors, 1):
                        cell = ws_horario.cell(row=ws_horario.max_row, column=col)
                        
                        if color_hex.startswith('#'):
                            color_hex = color_hex[1:]
                        
                        try:
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        except:
                            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                        
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.font = Font(color="ffffff")
                
                # Ajustar anchos de columnas
                ws_horario.column_dimensions['A'].width = 15
                for col in range(2, len(dias_orden) + 2):
                    ws_horario.column_dimensions[chr(64 + col)].width = 25
                
                # Ajustar altura de filas
                ws_horario.row_dimensions[1].height = 25
                for row in range(2, ws_horario.max_row + 1):
                    ws_horario.row_dimensions[row].height = 60
            
            # Eliminar la hoja de trabajo por defecto (Sheet) si se creó
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
                
            # Guardar el archivo
            wb.save(path)

            QMessageBox.information(self, "Exportar", f"Datos guardados correctamente en:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Fallo al exportar: {e}.\n¡Asegúrate de haber instalado 'openpyxl'!")

    def resaltar_boton(self, boton_activo):
        self.btnVistaHorario.setChecked(False)
        self.btnGestionDatos.setChecked(False)
        self.btnGenerarHorario.setChecked(False)
        boton_activo.setChecked(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ventana = MainApp()
    ventana.show()
    sys.exit(app.exec_())